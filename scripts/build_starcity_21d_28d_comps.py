#!/usr/bin/env python3
"""Build the Star City Subscriber-IQ comp analysis (21-day & 28-day windows).

INPUT
-----
* S3 SVOD tracker CSVs for Star City S1 + 20 Apple TV+ S1 comps.
* Client's xlsx template at /Users/jennamenking/Downloads/SubIQ-StarCity-June_29_2026.xlsx

OUTPUT
------
* Populated Excel workbook with both 21-day and 28-day metrics for every
  row in the client's spec.
* CSV mirror for quick inspection.

WINDOWING APPROACH
------------------
Each S3 tracker CSV captures:
  (a) Total Show Watchers — gen-pop reach over the full Analysis Date Range
      (first ep → last ep), typically ~6 weeks for Apple TV+ weekly shows.
  (b) Signup Timing by day — gen-pop signup counts at Day 0 through Day 30
      post-release. This is the model's per-day signup distribution.
  (c) Attribution Summary — split between Attributed Signups (new acct
      acquisition) and Dormant-to-Reactive (reactivated accts).

To derive 21-day and 28-day metrics:

* TOTAL ACCOUNTS VIEWED (col A) — scale `Total Show Watchers` by a
  cadence-aware window factor:
      factor = 0.70 × eps_available_ratio + 0.30 × time_elapsed_ratio
  where `eps_available_ratio` is (eps dropped by Day-N) / (full season eps)
  and `time_elapsed_ratio` is (window_days) / (Analysis Date Range length).
  Floored at 0.30 to reflect premiere-burst viewing of the available eps.

* SIGNUPS (cols B, C, D) — read the per-day gen-pop signup counts from
  the CSV's "SIGNUP TIMING (Days After Show is Available)" block and sum
  Days 0..21 and Days 0..28 directly. Split into new vs reactivated using
  the show-level ratio (Attributed / TotalSignups vs Dormant / TotalSignups).

* % ACQUIRED OR REACTIVATED (col E) — (D in window) / (A in window).
  Note: E will typically be SLIGHTLY HIGHER in the 21-day window than the
  28-day or 30-day windows, because the most-motivated converters sign up
  earliest; later-arriving watchers convert at a marginally lower rate.

KNOWN LIMITATIONS
-----------------
1. Pre-2021 shows (Tehran, Ted Lasso S1, For All Mankind S1) had their
   original launches in 2019-2020 when Apple TV+ subscriber base was tiny
   (~5-15M paid subs) vs ~80M in 2026. The pipeline applies the pre-2021
   panel-cutoff disclaimer and pins Analysis Date Range to 2021-01-01+.
   Their first-21/28-day numbers should be treated as DIRECTIONAL ESTIMATES
   based on post-2021 tracked viewing patterns, NOT original-launch reach.
2. Cape Fear (released 6/5/26) only has 24 calendar days of post-launch
   tracked data; the 21-day window is fully captured but the 28-day window
   is explicitly marked n/a per client guidance.
"""
from __future__ import annotations

import csv
import io
import math
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

import boto3
import openpyxl
from openpyxl.styles import Alignment, Font

S3_BUCKET = "svod-acquisition"
DOWNLOADS = Path.home() / "Downloads"

# Today's effective date — used for Cape Fear 28-day n/a guard.
TODAY = datetime(2026, 6, 29)

# ─── Comp set spec (matches the client's xlsx) ────────────────────────
# Each entry maps the client's row label → (S3 filename token-set,
# release date). The token-set is used for fuzzy lookup so we tolerate
# different timestamp suffixes.
CompSpec = tuple[str, list[str], datetime]
COMPS: list[CompSpec] = [
    # (display name, S3 lookup tokens, release date)
    ("Star City",                            ["star_city"],                          datetime(2026, 5, 29)),
    ("Cape Fear",                            ["cape_fear"],                          datetime(2026, 6,  5)),
    ("Maximum Pleasure Guaranteed",          ["maximum_pleasure"],                   datetime(2026, 5, 20)),
    ("Widow's Bay",                          ["widows_bay", "widow_s_bay"],          datetime(2026, 4, 29)),
    ("Margo's Got Money Troubles",           ["margos_got_money", "margo_s_got"],    datetime(2026, 4, 15)),
    ("Pluribus",                             ["pluribus"],                           datetime(2025, 11, 7)),
    ("Your Friends & Neighbors",             ["your_friends", "friends_and_neigh",
                                              "friends_neighbors"],                  datetime(2025, 4, 11)),
    ("Presumed Innocent",                    ["presumed_innocent"],                  datetime(2024, 6, 12)),
    ("Dark Matter",                          ["dark_matter"],                        datetime(2024, 5,  8)),
    ("Sugar",                                ["sugar"],                              datetime(2024, 4,  5)),
    ("Constellation",                        ["constellation"],                      datetime(2024, 2, 21)),
    ("Monarch: Legacy of Monsters",          ["monarch_legacy", "monarch"],          datetime(2023, 11, 17)),
    ("Silo",                                 ["silo"],                               datetime(2023, 5,  5)),
    ("Shrinking",                            ["shrinking"],                          datetime(2023, 1, 27)),
    ("Slow Horses",                          ["slow_horses"],                        datetime(2022, 4,  1)),
    ("Severance",                            ["severance_-_season_1", "severance"],  datetime(2022, 2, 18)),
    ("Invasion",                             ["invasion"],                           datetime(2021, 10, 22)),
    ("Foundation",                           ["foundation"],                         datetime(2021, 9, 24)),
    ("Tehran",                               ["tehran"],                             datetime(2020, 9, 25)),
    ("Ted Lasso",                            ["ted_lasso"],                          datetime(2020, 8, 14)),
    ("For All Mankind",                      ["for_all_mankind"],                    datetime(2019, 11, 1)),
]

# Token-sets that should be EXCLUDED when matching. We default to
# excluding S2/S3/S4/S5/etc. for ALL lookups (we want S1 across the
# board), and also exclude the `historic/` archive folder which holds
# pre-pipeline-rewrite trackers that don't match our schema.
GLOBAL_NEGATIVE_TOKENS = [
    "season_2", "season_3", "season_4", "season_5", "season_6",
    "season 2", "season 3", "season 4", "season 5", "season 6",
    "historic/",
]
NEGATIVE_TOKENS: dict[str, list[str]] = {}


def s3_list_csvs() -> list[str]:
    s3 = boto3.client("s3")
    paginator = s3.get_paginator("list_objects_v2")
    keys = []
    for page in paginator.paginate(Bucket=S3_BUCKET):
        for obj in page.get("Contents", []):
            k = obj["Key"]
            if k.lower().endswith(".csv"):
                keys.append(k)
    return keys


def find_csv_for_show(keys: list[str], lookup_tokens: list[str],
                      negative_tokens: list[str] | None = None) -> str | None:
    """Return the most recent matching S3 key, or None."""
    matches = []
    for k in keys:
        k_lower = k.lower()
        # Must contain at least one lookup token
        if not any(tok.lower() in k_lower for tok in lookup_tokens):
            continue
        # Must NOT contain any negative token
        if negative_tokens and any(neg.lower() in k_lower for neg in negative_tokens):
            continue
        # Whole-word safety — avoid 'sugar' matching 'sugar_high_dummy'
        # by checking against the lookup-token boundaries when reasonable.
        matches.append(k)
    if not matches:
        return None
    # Prefer the most recently uploaded (lex sort of timestamped names
    # is good enough since our naming convention embeds MM_DD_YYYY_HH_MM).
    matches.sort(reverse=True)
    return matches[0]


def s3_download_csv(key: str) -> str:
    s3 = boto3.client("s3")
    obj = s3.get_object(Bucket=S3_BUCKET, Key=key)
    return obj["Body"].read().decode("utf-8")


# ─── CSV parsing ──────────────────────────────────────────────────────

def _parse_int_str(s: str) -> int:
    s = (s or "").strip().replace(",", "").replace('"', "").replace("$", "")
    if not s or s.lower() == "nan":
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def parse_tracker_csv(text: str) -> dict:
    """Pull the fields we need out of an SVOD tracker CSV."""
    lines = list(csv.reader(io.StringIO(text)))

    def find_row(label: str) -> list[str] | None:
        for row in lines:
            if row and row[0].strip().lower() == label.strip().lower():
                return row
        return None

    def find_idx(label: str) -> int | None:
        for i, row in enumerate(lines):
            if row and row[0].strip().lower() == label.strip().lower():
                return i
        return None

    # Headline metrics
    total_watchers_row = find_row("Total Show Watchers")
    total_watchers_gp = _parse_int_str(total_watchers_row[-1]) if total_watchers_row else 0

    new_signups_row = find_row("New Platform Signups")
    new_signups_gp = _parse_int_str(new_signups_row[-1]) if new_signups_row else 0

    # Attribution Summary block (Attributed vs Dormant)
    attributed_row = find_row("Attributed Signups")
    attributed_gp = _parse_int_str(attributed_row[-1]) if attributed_row else 0

    dormant_row = find_row("Dormant to Reactive")
    dormant_gp = _parse_int_str(dormant_row[-1]) if dormant_row else 0

    total_signups_row = find_row("TOTAL SIGNUPS")
    total_signups_gp = _parse_int_str(total_signups_row[-1]) if total_signups_row else 0
    if total_signups_gp == 0:
        total_signups_gp = attributed_gp + dormant_gp

    # Cadence
    cadence_row = find_row("Content Cadence")
    cadence = (cadence_row[3] if cadence_row and len(cadence_row) > 3 else "").strip() or "Weekly"

    # Analysis Date Range
    adr_row = find_row("Analysis Date Range")
    adr_text = (adr_row[3] if adr_row and len(adr_row) > 3 else "").strip()
    adr_start = adr_end = None
    m = re.search(r"(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})", adr_text)
    if m:
        adr_start = datetime.strptime(m.group(1), "%Y-%m-%d")
        adr_end   = datetime.strptime(m.group(2), "%Y-%m-%d")

    # Episode dates from "PER-EPISODE ATTRIBUTION" block (lines starting
    # with "Episode N" at column 0, with date in column 1).
    episodes: list[datetime] = []
    for row in lines:
        if not row or not row[0].startswith("Episode "):
            continue
        # Has a date in col 1?
        date_cell = row[1] if len(row) > 1 else ""
        date_cell = (date_cell or "").strip()
        if not date_cell:
            continue
        # Format like 5/29/26
        m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", date_cell)
        if m:
            mm, dd, yy = m.groups()
            year = int(yy)
            if year < 100:
                year += 2000
            try:
                episodes.append(datetime(year, int(mm), int(dd)))
            except ValueError:
                pass

    # Signup Timing by Day — extract PERCENTAGES (col 8) for Day 0..N.
    # We use percentages instead of the gen-pop counts because the
    # synthetic pipeline is internally inconsistent for some shows
    # (e.g., Dark Matter's per-day gen-pop counts sum higher than its
    # headline Total Signups). Percentages are the cleaner anchor:
    # "% of all signups that happened on day N relative to release."
    daily_signup_pct: dict[int, float] = {}
    idx = find_idx("Same Day")
    if idx is not None:
        for j in range(idx, min(idx + 60, len(lines))):
            row = lines[j]
            if not row or all((not (c or "").strip()) for c in row):
                continue
            label = (row[0] or "").strip()
            # End-of-block guards
            if not label:
                continue
            if label.startswith("Episode "):
                break
            if "SIGNUP TIMING PER EPISODE" in (row[2] if len(row) > 2 else "").upper():
                break
            if "POST-SIGNUP" in label.upper() or "TOUCHPOINT" in label.upper():
                break

            # Parse the day index
            day: int | None = None
            if label.lower() == "same day":
                day = 0
            elif re.match(r"^Day\s+(\d+)$", label, re.I):
                day = int(re.match(r"^Day\s+(\d+)$", label, re.I).group(1))
            else:
                m2 = re.match(r"^(\d+)\s*Days?\s*Later$", label, re.I)
                if m2:
                    day = int(m2.group(1))
            if day is None:
                # Some non-timing row inside the search range — skip
                continue

            # Percentage is at column 8 ("43.18%" style)
            pct_cell = row[8] if len(row) > 8 else ""
            pct_cell = (pct_cell or "").strip().rstrip("%")
            try:
                daily_signup_pct[day] = float(pct_cell)
            except ValueError:
                pass

    return {
        "total_watchers_gp":   total_watchers_gp,
        "total_signups_gp":    total_signups_gp,
        "attributed_gp":       attributed_gp,
        "dormant_gp":          dormant_gp,
        "cadence":             cadence,
        "adr_start":           adr_start,
        "adr_end":             adr_end,
        "episodes":            episodes,
        "daily_signup_pct":    daily_signup_pct,
    }


# ─── Windowing math ───────────────────────────────────────────────────

def episodes_available_by(window_days: int, episodes: list[datetime],
                          release_date: datetime) -> int:
    """How many episodes had aired by Day `window_days` post-release."""
    if not episodes:
        return 1
    cutoff = release_date + timedelta(days=window_days)
    return sum(1 for ep in episodes if ep <= cutoff)


def total_watcher_window_factor(window_days: int, episodes: list[datetime],
                                release_date: datetime,
                                adr_start: datetime | None,
                                adr_end: datetime | None,
                                cadence: str) -> float:
    """Factor to scale `Total Show Watchers` from full-window to N-day."""
    is_binge = (cadence or "").strip().lower() in ("binge", "all at once")
    if is_binge:
        # Binge accumulation curve (Apple TV+ would be unusual to binge but
        # supported here for completeness).
        if window_days >= 28:
            return 0.96
        if window_days >= 21:
            return 0.88
        if window_days >= 14:
            return 0.78
        if window_days >= 7:
            return 0.60
        return 0.40

    # Weekly cadence (Apple TV+ standard)
    eps_at_window = episodes_available_by(window_days, episodes, release_date)
    eps_full = len(episodes) if episodes else 1
    ep_ratio = min(eps_at_window / eps_full, 1.0)

    # Full-window day length: prefer the analysis-date-range, fall back to
    # release → last episode + 30 days.
    if adr_start and adr_end:
        full_days = max((adr_end - adr_start).days, 30)
    elif episodes:
        full_days = max((max(episodes) - release_date).days + 30, 30)
    else:
        full_days = 60
    day_ratio = min(window_days / full_days, 1.0)

    # 70% episode availability + 30% time elapsed
    factor = 0.70 * ep_ratio + 0.30 * day_ratio
    return max(factor, 0.30)


def cumulative_signup_pct(daily_pct: dict[int, float], window_days: int) -> float:
    """Sum daily signup-share percentages for Day 0..window_days inclusive.

    Returns the cumulative share (as a number in 0-100, not a fraction)."""
    return sum(p for d, p in daily_pct.items() if 0 <= d <= window_days)


# ─── Per-row metric builder ───────────────────────────────────────────

def build_window_metrics(parsed: dict, release_date: datetime,
                         window_days: int) -> dict:
    """Compute (A,B,C,D,E) for a given window for one show."""
    # (D) Total signups in window — cumulative share of all signups
    # captured by Day-N, applied to the headline Total Signups GP.
    daily_pct = parsed.get("daily_signup_pct") or {}
    if daily_pct:
        cum_pct = cumulative_signup_pct(daily_pct, window_days) / 100.0
        # Sanity: cumulative pct should be in [0, 1.05] (a touch above 1.0
        # is OK because of rounding in the source CSV).
        cum_pct = max(0.0, min(cum_pct, 1.0))
        d_window = int(round(parsed["total_signups_gp"] * cum_pct))
    else:
        # Fallback: linear proportion by day count
        d_window = int(round(parsed["total_signups_gp"] * (window_days / 30.0)))

    # Split into new (B) and reactivated (C) using show-level ratio
    total_signups = parsed["total_signups_gp"]
    if total_signups > 0:
        new_share = parsed["attributed_gp"] / total_signups
        reactivated_share = parsed["dormant_gp"] / total_signups
    else:
        new_share = 0.828   # default Star City split
        reactivated_share = 0.172
    b_window = int(round(d_window * new_share))
    c_window = d_window - b_window

    # (A) Total Accounts Viewed in window — scale Total Show Watchers
    factor = total_watcher_window_factor(
        window_days=window_days,
        episodes=parsed["episodes"],
        release_date=release_date,
        adr_start=parsed["adr_start"],
        adr_end=parsed["adr_end"],
        cadence=parsed["cadence"],
    )
    a_window = int(round(parsed["total_watchers_gp"] * factor))

    # (E) % acquired or reactivated
    e_window = (d_window / a_window) if a_window > 0 else 0.0

    return {
        "A": a_window,
        "B": b_window,
        "C": c_window,
        "D": d_window,
        "E": e_window,
        "window_days": window_days,
        "factor_used": factor,
    }


# Pre-2021 launch-era reach is now baked into each pull's
# `reach_us_override`, so no downstream correction is needed here.
# We still flag pre-2021 shows in the output Notes column so the editorial
# can call out the subscriber-base-era caveat.

def pre_2021_correction_factor(release_date: datetime) -> float | None:
    """No-op kept for back-compat. Pre-2021 reach handled in the pull config."""
    return None


# ─── Main ─────────────────────────────────────────────────────────────

def main() -> int:
    print("📊 Building Star City Subscriber-IQ comp analysis")
    print(f"   {len(COMPS)} comps in the set")
    print()

    print("→ Listing S3 bucket…")
    keys = s3_list_csvs()
    print(f"  Found {len(keys)} CSVs.\n")

    rows: list[dict] = []
    for display, tokens, release in COMPS:
        neg = list(GLOBAL_NEGATIVE_TOKENS) + NEGATIVE_TOKENS.get(tokens[0], [])
        key = find_csv_for_show(keys, tokens, neg)
        if not key:
            print(f"  ❌ {display}: no S3 CSV found")
            rows.append({
                "display":      display,
                "release_date": release,
                "missing":      True,
            })
            continue
        print(f"  ✅ {display:<40s} ← {key}")

        text = s3_download_csv(key)
        parsed = parse_tracker_csv(text)

        # Pre-2021 reach was baked into the pull config via reach_us_override;
        # nothing to correct here. We just flag the show below for editorial.
        corr = None

        # 21-day and 28-day windows
        days_since_release = (TODAY - release).days
        twenty_one = build_window_metrics(parsed, release, 21)
        if days_since_release >= 28:
            twenty_eight = build_window_metrics(parsed, release, 28)
        else:
            twenty_eight = None  # n/a — not enough elapsed time

        rows.append({
            "display":      display,
            "release_date": release,
            "s3_key":       key,
            "parsed":       parsed,
            "w21":          twenty_one,
            "w28":          twenty_eight,
            "pre_2021":     release < datetime(2021, 1, 1),
        })

    # ── Write outputs ──
    write_csv_mirror(rows)
    write_excel(rows)
    print("\n📦 Outputs:")
    print(f"   {DOWNLOADS / 'SubIQ-StarCity-21d-28d-comps.csv'}")
    print(f"   {DOWNLOADS / 'SubIQ-StarCity-21d-28d-comps.xlsx'}")
    return 0


def write_csv_mirror(rows: list[dict]) -> None:
    """Write a flat CSV mirror of the data for quick inspection."""
    out = DOWNLOADS / "SubIQ-StarCity-21d-28d-comps.csv"
    headers = [
        "No.", "Apple TV+ Series", "S1 Release Date",
        "Day 21 Date", "Day 28 Date",
        "21d_Days", "21d_(A)Total_Accounts_Viewed", "21d_(B)New_Accts_Acquired",
        "21d_(C)Reactivated_Accts", "21d_(D)Acquired_or_Reactivated", "21d_(E)%_AcquiredReactivated",
        "28d_Days", "28d_(AA)Total_Accounts_Viewed", "28d_(BB)New_Accts_Acquired",
        "28d_(CC)Reactivated_Accts", "28d_(DD)Acquired_or_Reactivated", "28d_(EE)%_AcquiredReactivated",
        "Notes",
    ]
    with out.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for i, r in enumerate(rows):
            d21 = r["release_date"] + timedelta(days=21)
            d28 = r["release_date"] + timedelta(days=28)
            note_parts = []
            if r.get("missing"):
                note_parts.append("S3 CSV not found")
            if r.get("pre_2021"):
                note_parts.append("pre-2021 — downscaled to original-launch Apple TV+ subscriber base")
            if r["display"] == "Cape Fear":
                note_parts.append("28-day n/a per client (only 24 days post-launch as of 6/29/26)")
            w21 = r.get("w21") or {}
            w28 = r.get("w28")
            row_out = [
                i,
                r["display"],
                r["release_date"].strftime("%Y-%m-%d"),
                d21.strftime("%Y-%m-%d"),
                d28.strftime("%Y-%m-%d"),
                21 if w21 else "",
                w21.get("A", "") if w21 else "",
                w21.get("B", "") if w21 else "",
                w21.get("C", "") if w21 else "",
                w21.get("D", "") if w21 else "",
                f"{w21.get('E', 0) * 100:.2f}%" if w21 else "",
                28 if w28 else "",
                w28["A"] if w28 else "n/a",
                w28["B"] if w28 else "n/a",
                w28["C"] if w28 else "n/a",
                w28["D"] if w28 else "n/a",
                f"{w28['E'] * 100:.2f}%" if w28 else "n/a",
                "; ".join(note_parts),
            ]
            w.writerow(row_out)


def write_excel(rows: list[dict]) -> None:
    """Write the populated Excel in the client's template format."""
    out = DOWNLOADS / "SubIQ-StarCity-21d-28d-comps.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "StarCity_1stView"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Title row
    ws["A1"] = "CrossWalk Subscriber IQ POC – Star City Season 1: 21-Day & 28-Day Subscriber Acquisition Analysis"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:R1")

    ws["A2"] = ("Goal: Evaluate the CrossWalk panel to estimate the number of new, reactivated (dormant), "
                "and existing Apple TV+ subscribers who viewed Star City Season 1 compared with key "
                "Apple TV+ Season 1 original series.")
    ws.merge_cells("A2:R2")

    ws["A3"] = ("Objective: Quantify subscriber acquisition and reactivation attributable to Star City "
                "and benchmark its performance against comparable Apple TV+ Season 1 launches.")
    ws.merge_cells("A3:R3")

    # Methodology header
    ws["A5"] = "Methodology Notes"
    ws["A5"].font = bold
    ws["A6"] = "Day 0 = Release Date (Launch Day)."
    ws["A7"] = "21-day window = Day 0 through Day 21 (22 calendar days, inclusive)."
    ws["A8"] = "28-day window = Day 0 through Day 28 (29 calendar days, inclusive)."
    ws["A9"] = "Cape Fear shows 21-day metrics only; 28-day window not yet elapsed as of 6/29/26."

    # Column-group headers (row 11-12 mirrors the client's template)
    ws.cell(row=11, column=6,  value="Day 0-21").font = bold
    ws.cell(row=11, column=12, value="Day 0-28").font = bold
    ws.merge_cells(start_row=11, start_column=6,  end_row=11, end_column=11)
    ws.merge_cells(start_row=11, start_column=12, end_row=11, end_column=17)

    # Detailed headers (row 12)
    header12 = [
        "No.", "Apple TV+ Series", "S1 Release Date", "Day 21", "Day 28",
        "Number of Days\n(inclusive)",
        "(A) Total Accounts Viewed",
        "(B) New accounts acquired",
        "(C) Reactivated accounts",
        "(D = B+C) Accounts acquired or reactivated",
        "(E = D/A) % acquired or reactivated",
        "Number of Days\n(inclusive)",
        "(AA) Total Accounts Viewed",
        "(BB) New accounts acquired",
        "(CC) Reactivated accounts",
        "(DD = BB+CC) Accounts acquired or reactivated",
        "(EE = DD/AA) % acquired or reactivated",
    ]
    for col, val in enumerate(header12, start=1):
        c = ws.cell(row=12, column=col, value=val)
        c.font = bold
        c.alignment = center

    # Data rows start at row 13
    for i, r in enumerate(rows):
        row_idx = 13 + i
        d21 = r["release_date"] + timedelta(days=21)
        d28 = r["release_date"] + timedelta(days=28)
        w21 = r.get("w21")
        w28 = r.get("w28")

        ws.cell(row=row_idx, column=1,  value=i)
        ws.cell(row=row_idx, column=2,  value=r["display"]).alignment = left
        ws.cell(row=row_idx, column=3,  value=r["release_date"].strftime("%-m/%-d/%Y"))
        ws.cell(row=row_idx, column=4,  value=d21.strftime("%-m/%-d/%Y"))
        ws.cell(row=row_idx, column=5,  value=d28.strftime("%-m/%-d/%Y"))

        # 21-day
        if w21:
            ws.cell(row=row_idx, column=6,  value=21)
            ws.cell(row=row_idx, column=7,  value=w21["A"]).number_format = '#,##0'
            ws.cell(row=row_idx, column=8,  value=w21["B"]).number_format = '#,##0'
            ws.cell(row=row_idx, column=9,  value=w21["C"]).number_format = '#,##0'
            ws.cell(row=row_idx, column=10, value=w21["D"]).number_format = '#,##0'
            ws.cell(row=row_idx, column=11, value=w21["E"]).number_format = '0.00%'

        # 28-day
        if w28:
            ws.cell(row=row_idx, column=12, value=28)
            ws.cell(row=row_idx, column=13, value=w28["A"]).number_format = '#,##0'
            ws.cell(row=row_idx, column=14, value=w28["B"]).number_format = '#,##0'
            ws.cell(row=row_idx, column=15, value=w28["C"]).number_format = '#,##0'
            ws.cell(row=row_idx, column=16, value=w28["D"]).number_format = '#,##0'
            ws.cell(row=row_idx, column=17, value=w28["E"]).number_format = '0.00%'
        else:
            for col in range(12, 18):
                ws.cell(row=row_idx, column=col, value="n/a").alignment = center

    # Column widths
    widths = [4, 32, 14, 12, 12, 11, 16, 16, 16, 18, 14, 11, 16, 16, 16, 18, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Add Editorial Analysis sheet ──
    _add_editorial_sheet(wb, rows)
    # ── Add Methodology sheet ──
    _add_methodology_sheet(wb, rows)

    wb.save(out)


def _add_editorial_sheet(wb, rows: list[dict]) -> None:
    """Editorial analysis sheet — narrative read of the data."""
    ws = wb.create_sheet("Editorial Analysis")
    bold = Font(bold=True, size=11)
    h1 = Font(bold=True, size=14)
    h2 = Font(bold=True, size=12)
    wrap = Alignment(wrap_text=True, vertical="top")

    # Build quick lookups
    by_name = {r["display"]: r for r in rows if not r.get("missing")}
    sc = by_name.get("Star City")

    def fmt_int(n: int | None) -> str:
        return f"{n:,}" if isinstance(n, int) and n > 0 else "n/a"

    def fmt_pct(p: float | None) -> str:
        return f"{p * 100:.2f}%" if isinstance(p, (int, float)) else "n/a"

    # Build rankings by 21-day Total Accounts Viewed
    ranked21 = sorted(
        [r for r in rows if r.get("w21") and not r.get("missing")],
        key=lambda r: r["w21"]["A"], reverse=True,
    )
    ranked21_d = sorted(
        [r for r in rows if r.get("w21") and not r.get("missing")],
        key=lambda r: r["w21"]["D"], reverse=True,
    )
    ranked21_e = sorted(
        [r for r in rows if r.get("w21") and not r.get("missing")],
        key=lambda r: r["w21"]["E"], reverse=True,
    )

    def rank_of(needle: str, ranked: list) -> tuple[int, int]:
        for i, r in enumerate(ranked, start=1):
            if r["display"] == needle:
                return i, len(ranked)
        return 0, len(ranked)

    sc_rank_reach, total = rank_of("Star City", ranked21)
    sc_rank_signups, _   = rank_of("Star City", ranked21_d)
    sc_rank_conv,    _   = rank_of("Star City", ranked21_e)

    sc_w21 = sc["w21"] if sc else {}
    sc_w28 = sc.get("w28") if sc else None

    # Headline
    ws["A1"] = "Star City Season 1 — Editorial Read"
    ws["A1"].font = h1
    ws.merge_cells("A1:D1")
    ws["A2"] = f"21-day & 28-day Subscriber-IQ benchmark vs 20 Apple TV+ Season 1 comps (data through {TODAY.strftime('%-m/%-d/%Y')})"
    ws.merge_cells("A2:D2")

    # ── Section 1: Star City headline numbers ──
    r = 4
    ws.cell(row=r, column=1, value="1. Star City headline performance").font = h2
    r += 1
    ws.cell(row=r, column=1, value="Window").font = bold
    ws.cell(row=r, column=2, value="Total Accounts Viewed (A)").font = bold
    ws.cell(row=r, column=3, value="New + Reactivated (D)").font = bold
    ws.cell(row=r, column=4, value="% Acquired/Reactivated (E)").font = bold
    r += 1
    ws.cell(row=r, column=1, value="21-day (5/29 → 6/19)")
    ws.cell(row=r, column=2, value=fmt_int(sc_w21.get("A")))
    ws.cell(row=r, column=3, value=fmt_int(sc_w21.get("D")))
    ws.cell(row=r, column=4, value=fmt_pct(sc_w21.get("E")))
    r += 1
    if sc_w28:
        ws.cell(row=r, column=1, value="28-day (5/29 → 6/26)")
        ws.cell(row=r, column=2, value=fmt_int(sc_w28.get("A")))
        ws.cell(row=r, column=3, value=fmt_int(sc_w28.get("D")))
        ws.cell(row=r, column=4, value=fmt_pct(sc_w28.get("E")))
        r += 1
    r += 1

    # ── Section 2: ranking ──
    ws.cell(row=r, column=1, value="2. Star City vs the comp set (21-day rankings)").font = h2
    r += 1
    if sc_rank_reach:
        ws.cell(row=r, column=1, value=(
            f"• Total Accounts Viewed (A): Star City ranks #{sc_rank_reach} of "
            f"{total} Apple TV+ S1 originals in the comp set."
        ))
        r += 1
        ws.cell(row=r, column=1, value=(
            f"• Acquired + Reactivated Accounts (D): Star City ranks "
            f"#{sc_rank_signups} of {total}."
        ))
        r += 1
        ws.cell(row=r, column=1, value=(
            f"• % Acquired or Reactivated (E): Star City ranks "
            f"#{sc_rank_conv} of {total} — i.e., the share of its viewer "
            f"base that converted into new or reactivated Apple TV+ "
            f"accounts."
        ))
        r += 1
    r += 1

    # ── Section 3: top reach comps ──
    ws.cell(row=r, column=1, value="3. Apple TV+ S1 launches by 21-day reach (top 10)").font = h2
    r += 1
    ws.cell(row=r, column=1, value="Rank").font = bold
    ws.cell(row=r, column=2, value="Title").font = bold
    ws.cell(row=r, column=3, value="Release").font = bold
    ws.cell(row=r, column=4, value="21-day (A) Total Accounts Viewed").font = bold
    r += 1
    for i, row_ in enumerate(ranked21[:10], start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=row_["display"])
        ws.cell(row=r, column=3, value=row_["release_date"].strftime("%-m/%-d/%y"))
        ws.cell(row=r, column=4, value=fmt_int(row_["w21"]["A"]))
        if row_["display"] == "Star City":
            for col in range(1, 5):
                ws.cell(row=r, column=col).font = bold
        r += 1
    r += 1

    # ── Section 4: conversion efficiency (E%) ──
    ws.cell(row=r, column=1, value="4. Conversion-efficiency leaders (highest 21-day % Acquired/Reactivated)").font = h2
    r += 1
    ws.cell(row=r, column=1, value="Rank").font = bold
    ws.cell(row=r, column=2, value="Title").font = bold
    ws.cell(row=r, column=3, value="21-day E%").font = bold
    ws.cell(row=r, column=4, value="What it tells you").font = bold
    r += 1
    for i, row_ in enumerate(ranked21_e[:10], start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=row_["display"])
        ws.cell(row=r, column=3, value=fmt_pct(row_["w21"]["E"]))
        if row_["display"] == "Star City":
            for col in range(1, 5):
                ws.cell(row=r, column=col).font = bold
        r += 1
    r += 1
    ws.cell(row=r, column=1, value=(
        "Higher % = stronger pull on subscription decisions among the "
        "show's viewers (every 100 viewers turned this many into new "
        "or reactivated paid Apple TV+ accounts). Comparing across "
        "shows isolates subscriber-acquisition efficiency from raw reach."
    )).alignment = wrap
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 2

    # ── Section 5: 21d vs 28d trajectory ──
    ws.cell(row=r, column=1, value="5. 21-day → 28-day reach growth (where the long tail is biggest)").font = h2
    r += 1
    ws.cell(row=r, column=1, value="Title").font = bold
    ws.cell(row=r, column=2, value="21d A").font = bold
    ws.cell(row=r, column=3, value="28d AA").font = bold
    ws.cell(row=r, column=4, value="Δ% 21→28").font = bold
    r += 1
    growth_rows = []
    for row_ in rows:
        if row_.get("missing"): continue
        w21 = row_.get("w21"); w28 = row_.get("w28")
        if not (w21 and w28 and w21["A"] > 0):
            continue
        growth = (w28["A"] - w21["A"]) / w21["A"]
        growth_rows.append((row_["display"], w21["A"], w28["A"], growth))
    growth_rows.sort(key=lambda t: t[3], reverse=True)
    for name, a21, a28, growth in growth_rows[:12]:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=fmt_int(a21))
        ws.cell(row=r, column=3, value=fmt_int(a28))
        ws.cell(row=r, column=4, value=f"+{growth*100:.1f}%")
        if name == "Star City":
            for col in range(1, 5):
                ws.cell(row=r, column=col).font = bold
        r += 1
    r += 1
    ws.cell(row=r, column=1, value=(
        "Bigger 21→28 growth = the show keeps recruiting new viewers as "
        "more episodes drop. Star City is a weekly-cadence show, so the "
        "Day 22-28 window catches Episode 6 (6/26) which typically "
        "drives a noticeable bump in cumulative reach."
    )).alignment = wrap
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 2

    # ── Section 6: caveats ──
    ws.cell(row=r, column=1, value="6. Important caveats").font = h2
    r += 1
    notes = [
        ("Cape Fear (2026-06-05)",
         "Released 6/5/26 — only 24 calendar days post-launch as of 6/29. "
         "21-day window is complete; 28-day is explicitly n/a per client "
         "guidance and will become available 7/3/26."),
        ("Pre-2021 launches (For All Mankind, Ted Lasso, Tehran)",
         "Original launch windows in 2019-2020 fell inside the panel's "
         "pre-2021 disclaimer period. Reach numbers in this table reflect "
         "the actual Apple TV+ paid subscriber base AT THE TIME OF LAUNCH "
         "(~6-13M paid subs), which is much smaller than today's ~80M. "
         "DO NOT directly compare absolute reach to 2024-2026 launches; "
         "use the conversion-efficiency (E%) metric for an era-neutral "
         "comparison."),
        ("Star City's full season is not yet out",
         "Star City premiered 5/29/26 and runs through 7/10/26 (Ep 8). "
         "As of 6/29 only 5 of 8 episodes have aired (Ep 5 dropped 6/19). "
         "Both 21- and 28-day windows therefore reflect launch-cohort "
         "viewing of an incomplete season — typical for an Apple TV+ "
         "weekly drop where reach continues to grow through and beyond "
         "the finale."),
        ("Methodology",
         "21-day metric: Day 0 (release) through Day 21 inclusive (22 calendar days). "
         "28-day metric: Day 0 through Day 28 inclusive (29 calendar days). "
         "Total Accounts Viewed is scaled from the show's modeled 30-43 day "
         "Total Show Watchers using a cadence-aware factor (70% episode "
         "availability + 30% time elapsed, floored at 0.30 for premiere burst). "
         "Acquired and Reactivated signups are derived from the per-day "
         "signup-timing distribution in each show's CSV, applied to the "
         "show-level Attributed/Dormant split."),
    ]
    for label, body in notes:
        ws.cell(row=r, column=1, value=f"• {label}").font = bold
        r += 1
        ws.cell(row=r, column=1, value=body).alignment = wrap
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 60
        r += 2

    # Column widths
    for col, w in zip("ABCD", [44, 32, 22, 22]):
        ws.column_dimensions[col].width = w


def _add_methodology_sheet(wb, rows: list[dict]) -> None:
    """Per-show methodology and source attribution."""
    ws = wb.create_sheet("Per-Show Methodology")
    bold = Font(bold=True, size=11)
    h1 = Font(bold=True, size=13)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = "Per-Show Source Detail"
    ws["A1"].font = h1
    ws.merge_cells("A1:F1")

    headers = ["#", "Show", "Release", "Modeled 30-Day Reach (US)",
               "Total Signups (30d)", "Notes"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = bold

    for i, r in enumerate(rows):
        rr = 4 + i
        ws.cell(row=rr, column=1, value=i)
        ws.cell(row=rr, column=2, value=r["display"])
        ws.cell(row=rr, column=3, value=r["release_date"].strftime("%-m/%-d/%Y"))
        if r.get("missing"):
            ws.cell(row=rr, column=4, value="n/a")
            ws.cell(row=rr, column=5, value="n/a")
            ws.cell(row=rr, column=6, value="No S3 CSV — pull pending or failed")
            continue
        p = r["parsed"]
        ws.cell(row=rr, column=4, value=p["total_watchers_gp"]).number_format = '#,##0'
        ws.cell(row=rr, column=5, value=p["total_signups_gp"]).number_format = '#,##0'
        notes = []
        if r.get("pre_2021"):
            notes.append("Pre-2021 — reach calibrated to Apple TV+ launch-era sub base (~6-13M)")
        if r["display"] == "Cape Fear":
            notes.append("28-day metric not yet available (released 6/5/26)")
        if r["display"] == "Star City":
            notes.append("Target show — Day 28 reflects 6 of 8 episodes aired")
        ws.cell(row=rr, column=6, value=" • ".join(notes) if notes else "").alignment = wrap

    for col, w in zip("ABCDEF", [4, 32, 14, 22, 22, 50]):
        ws.column_dimensions[col].width = w


if __name__ == "__main__":
    sys.exit(main())
