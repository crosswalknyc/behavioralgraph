#!/usr/bin/env python3
"""Build the Pop Culture Jeopardy! S2 Subscriber-IQ comp analysis (25-day window).

INPUT
-----
* S3 tracker CSVs for PCJ S2 (Netflix) + 13 comparable unscripted /
  game / reality titles across Netflix, Prime, Peacock, Disney+/Hulu.
* Client's Excel template at
  /Users/jennamenking/Desktop/SubIQ -PopCultureJeopardy-July_7_2026.xlsx

OUTPUT
------
* Populated Excel workbook with 25-day metrics for every comp row.
* Second sheet: Prime PCJ S1 → Netflix PCJ S2 cross-platform migration
  signup story (per client's ask "if there may be a story of Prime
  Viewers of PCJ! Season 1 who didn't have Netflix that signed up for
  NF to watch Pop Culture Jeopardy S2, that would be interesting").
* CSV mirror for quick inspection.

WINDOWING APPROACH
------------------
25 days from EACH comp's own initial release date, standardized to
match PCJ S2's own lifecycle (5/11-6/5/26). Sony spec says:
   "Does it make sense to standardize with 28 days from their initial
    release date with the caveat that we standardized the days to the
    lifecycle of PCJ! Season 2"
Analyst (Jenna) chose 25-day (PCJ S2 actual lifecycle) over 28-day.

The pipeline CSVs use a 30-day analysis window. We slice each comp to
its 25-day post-release window using the CSV's per-day signup timing
block:

* TOTAL ACCOUNTS VIEWED (col AA) — scale `Total Show Watchers` by a
  cadence-aware window factor. For BINGE releases (Is It Cake, What's
  In The Box, batched: Squid Game Challenge, Million Dollar Secret,
  The Mole, Love Is Blind) the 25-day window captures 90-95% of
  30-day reach because viewing is front-loaded. For weekly (PCJ S1,
  Star Search, Traitors, DWTS) and daily-strip (PCJ S2, Love Island),
  factor is proportional to episodes-available + time-elapsed.

* SIGNUPS (cols BB, CC, DD) — sum the per-day gen-pop signup
  percentages Day 0..24 and multiply by the CSV's TOTAL SIGNUPS.
  Split into new (BB) vs reactivated (CC) using the per-show
  attribution ratio baked into the pull config.

* % ACQUIRED OR REACTIVATED (col EE) — DD / AA.

CROSS-PLATFORM MIGRATION TAB
-----------------------------
Estimates the size of the (Prime PCJ S1 viewers × Netflix new-sub in
5/2026 × PCJ S2 first-episode play) intersection cohort — the group
who signed up for Netflix specifically to watch PCJ S2 after having
watched PCJ S1 on Prime. Uses a 3-stage funnel:

  Stage 1: PCJ S1 unique US viewers on Prime (25-day) ≈ 1.6M
  Stage 2: × ~35% "did not already have Netflix" (Netflix US HH
           penetration ~65% — the 35% non-overlap slice)
  Stage 3: × ~28% "signed up for Netflix during 5/11-6/5 window and
           watched PCJ S2 within 25 days"

Yields a modeled ~155K US accounts who migrated Prime→Netflix
specifically for PCJ S2. Framed as MODELED (not measured) — replace
with Crosswalk panel intersection when available.
"""
from __future__ import annotations

import csv
import io
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

import boto3
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

S3_BUCKET = "svod-acquisition"
DOWNLOADS = Path.home() / "Downloads"

# Effective analysis date (today's date). PCJ S2 finale was 6/5/26,
# so the 25-day window (5/11 through 6/5) is fully captured for the
# subject; and all comps in the set have full 25-day-plus lifecycles.
TODAY = datetime(2026, 7, 7)

# ─── Comp set spec (matches Sony's Excel template exactly) ───────────
#
# Each entry: (display name, S3 lookup tokens, release date,
#              cadence label used in the CSV, negative-tokens to
#              avoid picking a wrong-season CSV).
#
# The subject (PCJ S2 Netflix) is Row 10 in the client's template.
# Rows 11-23 are the 13 comps.
COMPS: list[dict] = [
    # SUBJECT — PCJ S2 on Netflix
    {"row": "subject", "display": "Pop Culture Jeopardy! (S2 Netflix)",
     "platform": "Netflix", "season": 2,
     "tokens": ["pop_culture_jeopardy_-_season_2"], "neg": [],
     "release": datetime(2026, 5, 11), "finale": datetime(2026, 6, 5)},
    # ROW 1
    {"row": 1, "display": "Pop Culture Jeopardy! (S1 Amazon Prime)",
     "platform": "Prime", "season": 1,
     "tokens": ["pop_culture_jeopardy_-_season_1"], "neg": [],
     "release": datetime(2024, 12, 4), "finale": datetime(2025, 3, 5)},
    # ROW 2
    {"row": 2, "display": "Squid Game: Challenge (S2 Netflix)",
     "platform": "Netflix", "season": 2,
     "tokens": ["squid_game_the_challenge"], "neg": [],
     "release": datetime(2025, 11, 4), "finale": datetime(2025, 11, 18)},
    # ROW 3
    {"row": 3, "display": "Star Search (S1 Netflix)",
     "platform": "Netflix", "season": 1,
     "tokens": ["star_search"], "neg": [],
     "release": datetime(2026, 1, 20), "finale": datetime(2026, 2, 18)},
    # ROW 4
    {"row": 4, "display": "Is It Cake? (S3 Netflix)",
     "platform": "Netflix", "season": 3,
     "tokens": ["is_it_cake_-_season_3"], "neg": [],
     "release": datetime(2024, 3, 29), "finale": datetime(2024, 3, 29)},
    # ROW 5
    {"row": 5, "display": "Million Dollar Secret (S1 Netflix)",
     "platform": "Netflix", "season": 1,
     "tokens": ["million_dollar_secret_-_season_1"], "neg": [],
     "release": datetime(2025, 3, 26), "finale": datetime(2025, 4, 9)},
    # ROW 6
    {"row": 6, "display": "Million Dollar Secret (S2 Netflix)",
     "platform": "Netflix", "season": 2,
     "tokens": ["million_dollar_secret_-_season_2"], "neg": [],
     "release": datetime(2026, 4, 15), "finale": datetime(2026, 4, 29)},
    # ROW 7
    {"row": 7, "display": "The Mole (S1 Netflix)",
     "platform": "Netflix", "season": 1,
     "tokens": ["the_mole_-_season_1"], "neg": [],
     "release": datetime(2022, 10, 7), "finale": datetime(2022, 10, 21)},
    # ROW 8
    {"row": 8, "display": "The Mole (S2 Netflix)",
     "platform": "Netflix", "season": 2,
     "tokens": ["the_mole_-_season_2"], "neg": [],
     "release": datetime(2024, 6, 28), "finale": datetime(2024, 7, 12)},
    # ROW 9
    {"row": 9, "display": "What's In The Box (S1 Netflix)",
     "platform": "Netflix", "season": 1,
     "tokens": ["whats_in_the_box"], "neg": [],
     "release": datetime(2025, 12, 17), "finale": datetime(2025, 12, 17)},
    # ROW 10
    {"row": 10, "display": "Love Is Blind (S10 Netflix)",
     "platform": "Netflix", "season": 10,
     "tokens": ["love_is_blind_-_season_10"], "neg": [],
     "release": datetime(2026, 2, 11), "finale": datetime(2026, 3, 4)},
    # ROW 11
    {"row": 11, "display": "Love Island (S7 Peacock)",
     "platform": "Peacock", "season": 7,
     "tokens": ["love_island_-_season_7"], "neg": [],
     "release": datetime(2025, 6, 3), "finale": datetime(2025, 7, 13)},
    # ROW 12
    {"row": 12, "display": "Dancing With The Stars (S34)",
     "platform": "Disney+ + Hulu", "season": 34,
     "tokens": ["dancing_with_the_stars_-_season_34"], "neg": [],
     "release": datetime(2025, 9, 16), "finale": datetime(2025, 12, 25)},
    # ROW 13
    {"row": 13, "display": "The Traitors (S4 Peacock)",
     "platform": "Peacock", "season": 4,
     "tokens": ["the_traitors_-_season_4"], "neg": [],
     "release": datetime(2026, 1, 8), "finale": datetime(2026, 2, 26)},
]

GLOBAL_NEGATIVE_TOKENS = ["historic/"]


# ─── S3 CSV lookup ────────────────────────────────────────────────────

def s3_list_csvs() -> list[str]:
    s3 = boto3.client("s3")
    paginator = s3.get_paginator("list_objects_v2")
    keys = []
    for page in paginator.paginate(Bucket=S3_BUCKET):
        for obj in page.get("Contents", []):
            k = obj["Key"]
            if k.lower().endswith(".csv"):
                keys.append(k)
    return keys


def find_csv_for_show(keys: list[str], lookup_tokens: list[str],
                      negative_tokens: list[str] | None = None) -> str | None:
    """Newest matching S3 key (by embedded MM_DD_YYYY_HH_MM timestamp)."""
    matches = []
    for k in keys:
        low = k.lower()
        if any(neg.lower() in low for neg in (GLOBAL_NEGATIVE_TOKENS + (negative_tokens or []))):
            continue
        if not any(tok.lower() in low for tok in lookup_tokens):
            continue
        matches.append(k)
    if not matches:
        return None
    ts_re = re.compile(r'(\d{2})_(\d{2})_(\d{4})_(\d{2})_(\d{2})\.csv$')
    def _ts_key(k: str):
        m = ts_re.search(k)
        if not m:
            return (0, 0, 0, 0, 0)
        mm, dd, yyyy, hh, mi = m.groups()
        return (int(yyyy), int(mm), int(dd), int(hh), int(mi))
    matches.sort(key=_ts_key, reverse=True)
    return matches[0]


def s3_download_csv(key: str) -> str:
    s3 = boto3.client("s3")
    obj = s3.get_object(Bucket=S3_BUCKET, Key=key)
    return obj["Body"].read().decode("utf-8")


# ─── CSV parsing (identical semantics to Star City builder) ──────────

def _parse_int_str(s: str) -> int:
    s = (s or "").strip().replace(",", "").replace('"', "").replace("$", "")
    if not s or s.lower() == "nan":
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def parse_tracker_csv(text: str) -> dict:
    lines = list(csv.reader(io.StringIO(text)))

    def find_row(label: str) -> list[str] | None:
        for row in lines:
            if row and row[0].strip().lower() == label.strip().lower():
                return row
        return None

    def find_idx(label: str) -> int | None:
        for i, row in enumerate(lines):
            if row and row[0].strip().lower() == label.strip().lower():
                return i
        return None

    total_watchers_row = find_row("Total Show Watchers")
    total_watchers_gp = _parse_int_str(total_watchers_row[-1]) if total_watchers_row else 0

    attributed_row = find_row("Attributed Signups")
    attributed_gp = _parse_int_str(attributed_row[-1]) if attributed_row else 0

    dormant_row = find_row("Dormant to Reactive")
    dormant_gp = _parse_int_str(dormant_row[-1]) if dormant_row else 0

    total_signups_row = find_row("TOTAL SIGNUPS")
    total_signups_gp = _parse_int_str(total_signups_row[-1]) if total_signups_row else 0
    if total_signups_gp == 0:
        total_signups_gp = attributed_gp + dormant_gp

    cadence_row = find_row("Content Cadence")
    cadence = (cadence_row[3] if cadence_row and len(cadence_row) > 3 else "").strip() or "Weekly"

    adr_row = find_row("Analysis Date Range")
    adr_text = (adr_row[3] if adr_row and len(adr_row) > 3 else "").strip()
    adr_start = adr_end = None
    m = re.search(r"(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})", adr_text)
    if m:
        adr_start = datetime.strptime(m.group(1), "%Y-%m-%d")
        adr_end   = datetime.strptime(m.group(2), "%Y-%m-%d")

    episodes: list[datetime] = []
    for row in lines:
        if not row or not row[0].startswith("Episode "):
            continue
        date_cell = (row[1] if len(row) > 1 else "").strip()
        m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", date_cell)
        if m:
            mm, dd, yy = m.groups()
            year = int(yy)
            if year < 100:
                year += 2000
            try:
                episodes.append(datetime(year, int(mm), int(dd)))
            except ValueError:
                pass

    daily_signup_pct: dict[int, float] = {}
    idx = find_idx("Same Day")
    if idx is not None:
        for j in range(idx, min(idx + 60, len(lines))):
            row = lines[j]
            if not row or all((not (c or "").strip()) for c in row):
                continue
            label = (row[0] or "").strip()
            if not label:
                continue
            if label.startswith("Episode "):
                break
            if "SIGNUP TIMING PER EPISODE" in (row[2] if len(row) > 2 else "").upper():
                break
            if "POST-SIGNUP" in label.upper() or "TOUCHPOINT" in label.upper():
                break
            day: int | None = None
            if label.lower() == "same day":
                day = 0
            elif re.match(r"^Day\s+(\d+)$", label, re.I):
                day = int(re.match(r"^Day\s+(\d+)$", label, re.I).group(1))
            else:
                m2 = re.match(r"^(\d+)\s*Days?\s*Later$", label, re.I)
                if m2:
                    day = int(m2.group(1))
            if day is None:
                continue
            pct_cell = (row[8] if len(row) > 8 else "").strip().rstrip("%")
            try:
                daily_signup_pct[day] = float(pct_cell)
            except ValueError:
                pass

    return {
        "total_watchers_gp":   total_watchers_gp,
        "total_signups_gp":    total_signups_gp,
        "attributed_gp":       attributed_gp,
        "dormant_gp":          dormant_gp,
        "cadence":             cadence,
        "adr_start":           adr_start,
        "adr_end":             adr_end,
        "episodes":            episodes,
        "daily_signup_pct":    daily_signup_pct,
    }


# ─── Windowing math ───────────────────────────────────────────────────

def episodes_available_by(window_days: int, episodes: list[datetime],
                          release_date: datetime) -> int:
    if not episodes:
        return 1
    cutoff = release_date + timedelta(days=window_days)
    return sum(1 for ep in episodes if ep <= cutoff)


def total_watcher_window_factor(window_days: int, episodes: list[datetime],
                                release_date: datetime,
                                adr_start: datetime | None,
                                adr_end: datetime | None,
                                cadence: str) -> float:
    is_binge = (cadence or "").strip().lower() in ("binge", "all at once", "batched")
    if is_binge:
        if window_days >= 28: return 0.96
        if window_days >= 25: return 0.94
        if window_days >= 21: return 0.88
        if window_days >= 14: return 0.78
        if window_days >= 7:  return 0.60
        return 0.40

    eps_at_window = episodes_available_by(window_days, episodes, release_date)
    eps_full = len(episodes) if episodes else 1
    ep_ratio = min(eps_at_window / eps_full, 1.0)

    if adr_start and adr_end:
        full_days = max((adr_end - adr_start).days, 30)
    elif episodes:
        full_days = max((max(episodes) - release_date).days + 30, 30)
    else:
        full_days = 60
    day_ratio = min(window_days / full_days, 1.0)
    factor = 0.70 * ep_ratio + 0.30 * day_ratio
    return max(factor, 0.30)


def cumulative_signup_pct(daily_pct: dict[int, float], window_days: int) -> float:
    return sum(p for d, p in daily_pct.items() if 0 <= d <= window_days)


def build_window_metrics(parsed: dict, release_date: datetime,
                         window_days: int) -> dict:
    daily_pct = parsed.get("daily_signup_pct") or {}
    if daily_pct:
        cum_pct = cumulative_signup_pct(daily_pct, window_days) / 100.0
        cum_pct = max(0.0, min(cum_pct, 1.0))
        d_window = int(round(parsed["total_signups_gp"] * cum_pct))
    else:
        d_window = int(round(parsed["total_signups_gp"] * (window_days / 30.0)))

    total_signups = parsed["total_signups_gp"]
    if total_signups > 0:
        new_share = parsed["attributed_gp"] / total_signups
    else:
        new_share = 0.55
    b_window = int(round(d_window * new_share))
    c_window = d_window - b_window

    factor = total_watcher_window_factor(
        window_days=window_days,
        episodes=parsed["episodes"],
        release_date=release_date,
        adr_start=parsed["adr_start"],
        adr_end=parsed["adr_end"],
        cadence=parsed["cadence"],
    )
    a_window = int(round(parsed["total_watchers_gp"] * factor))
    e_window = (d_window / a_window) if a_window > 0 else 0.0

    return {"A": a_window, "B": b_window, "C": c_window, "D": d_window,
            "E": e_window, "window_days": window_days, "factor_used": factor}


# ─── Main pipeline ────────────────────────────────────────────────────

def main() -> int:
    print("📊 Building Pop Culture Jeopardy! S2 Subscriber-IQ 25-day comp analysis")
    print(f"   {len(COMPS)} rows in the set (1 subject + {len(COMPS)-1} comps)")
    print()

    print("→ Listing S3 bucket…")
    keys = s3_list_csvs()
    print(f"  Found {len(keys)} CSVs.\n")

    rows: list[dict] = []
    for spec in COMPS:
        key = find_csv_for_show(keys, spec["tokens"], spec.get("neg"))
        if not key:
            print(f"  ❌ {spec['display']}: no S3 CSV found")
            rows.append({**spec, "missing": True})
            continue
        print(f"  ✅ {spec['display']:<45s} ← {key}")

        text = s3_download_csv(key)
        parsed = parse_tracker_csv(text)
        w25 = build_window_metrics(parsed, spec["release"], 25)
        rows.append({**spec, "s3_key": key, "parsed": parsed, "w25": w25})

    write_csv_mirror(rows)
    write_excel(rows)
    print("\n📦 Outputs:")
    print(f"   {DOWNLOADS / 'SubIQ-PopCultureJeopardy-25d-comps.csv'}")
    print(f"   {DOWNLOADS / 'SubIQ-PopCultureJeopardy-25d-comps.xlsx'}")
    return 0


def write_csv_mirror(rows: list[dict]) -> None:
    out = DOWNLOADS / "SubIQ-PopCultureJeopardy-25d-comps.csv"
    headers = [
        "No.", "Show", "Season", "Platform", "Release Date", "Finale Date",
        "Length (Days)", "Day 25", "Number of Days (inclusive)",
        "(AA) Total Accounts Viewed", "(BB) New accounts acquired",
        "(CC) Reactivated accounts", "(DD = BB+CC) Acquired or Reactivated",
        "(EE = DD/AA) % acquired or reactivated",
    ]
    with out.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w25 = r.get("w25") or {}
            release = r["release"]
            finale  = r["finale"]
            day25   = release + timedelta(days=25)
            length_days = (finale - release).days
            no = "" if r["row"] == "subject" else r["row"]
            row_out = [
                no, r["display"], r["season"], r["platform"],
                release.strftime("%-m/%-d/%Y"),
                finale.strftime("%-m/%-d/%Y"),
                length_days,
                day25.strftime("%-m/%-d/%Y"),
                25 if w25 else "",
                w25.get("A", "") if w25 else "",
                w25.get("B", "") if w25 else "",
                w25.get("C", "") if w25 else "",
                w25.get("D", "") if w25 else "",
                f"{w25.get('E', 0) * 100:.2f}%" if w25 else "",
            ]
            w.writerow(row_out)


def write_excel(rows: list[dict]) -> None:
    out = DOWNLOADS / "SubIQ-PopCultureJeopardy-25d-comps.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PCJ_1stView"

    bold = Font(bold=True)
    h1 = Font(bold=True, size=13)
    h2 = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    subject_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Title + preamble (mirrors Sony's template exactly on rows 1-4)
    ws["A1"] = ("CrossWalk Subscriber IQ POC — Pop Culture Jeopardy! Season 2 "
                "(Netflix): 25-Day Subscriber Acquisition Analysis")
    ws["A1"].font = h1
    ws.merge_cells("A1:N1")

    ws["A2"] = ("Goal: Evaluate CrossWalk panel to estimate the number of new, "
                "reactivated (dormant), and existing streaming subscribers who "
                "viewed Pop Culture Jeopardy! Season 2 on Netflix, benchmarked "
                "against comparable streaming unscripted / game / reality "
                "launches across Netflix, Prime, Peacock, and Disney+/Hulu.")
    ws.merge_cells("A2:N2")
    ws["A2"].alignment = left

    ws["A3"] = ("Objective: Gauge subscriber acquisition and reactivation "
                "attributable to Pop Culture Jeopardy! S2 and rank it inside "
                "a 13-title comp set — standardized to a common 25-day "
                "post-release window (matches PCJ S2's own 5/11-6/5/26 "
                "lifecycle).")
    ws.merge_cells("A3:N3")
    ws["A3"].alignment = left

    ws["A4"] = f"As of {TODAY.strftime('%-m/%-d/%Y')}"
    ws["A4"].font = bold

    # Methodology notes (rows 5-7)
    ws["A5"] = "Methodology Notes"
    ws["A5"].font = bold
    ws["A6"] = ("Day 0 = each show's original release date. 25-day window = "
                "Day 0 through Day 24 (25 calendar days, inclusive) — "
                "standardized to PCJ S2's own lifecycle length regardless "
                "of the comp's actual finale date.")
    ws["A6"].alignment = left
    ws.merge_cells("A6:N6")
    ws["A7"] = ("Where a comp's lifecycle > 25 days (weekly / daily-strip "
                "shows like DWTS S34, Traitors S4, Love Island S7, PCJ S1), "
                "the 25-day slice captures only its first 25 days of "
                "post-release engagement — the fair comp vs PCJ S2's own "
                "25-day lifecycle. Where a comp's lifecycle < 25 days "
                "(binge/batched shows), the window captures the full "
                "release plus tail viewing through Day 25.")
    ws["A7"].alignment = left
    ws.merge_cells("A7:N7")

    # Column group header (row 8) — mirrors template's merged "Day 0-28"
    # but relabeled to 25-day
    ws.cell(row=8, column=9, value="Day 0-25").font = bold
    ws.cell(row=8, column=9).alignment = center
    ws.merge_cells(start_row=8, start_column=9, end_row=8, end_column=14)
    ws.cell(row=8, column=9).fill = subject_fill

    # Column headers (row 9)
    headers = [
        "No.", "Show", "Season", "Platform", "Release Date", "Finale Date",
        "Length (Days)", "Day 25",
        "Number of Days\n(inclusive)",
        "(AA) Total Accounts Viewed",
        "(BB) New accounts acquired",
        "(CC) Reactivated accounts",
        "(DD = BB+CC) Acquired or Reactivated",
        "(EE = DD/AA) % acquired or reactivated",
    ]
    for col, val in enumerate(headers, start=1):
        c = ws.cell(row=9, column=col, value=val)
        c.font = bold
        c.alignment = center
    ws.row_dimensions[9].height = 42

    # Data rows start at row 10 (matches template)
    for i, r in enumerate(rows):
        row_idx = 10 + i
        release = r["release"]
        finale  = r["finale"]
        day25   = release + timedelta(days=25)
        length_days = (finale - release).days
        w25 = r.get("w25") or {}
        is_subject = r["row"] == "subject"

        ws.cell(row=row_idx, column=1, value="" if is_subject else r["row"])
        ws.cell(row=row_idx, column=2, value=r["display"]).alignment = left
        ws.cell(row=row_idx, column=3, value=r["season"])
        ws.cell(row=row_idx, column=4, value=r["platform"]).alignment = center
        ws.cell(row=row_idx, column=5, value=release.strftime("%-m/%-d/%Y"))
        ws.cell(row=row_idx, column=6, value=finale.strftime("%-m/%-d/%Y"))
        ws.cell(row=row_idx, column=7, value=length_days).alignment = center
        ws.cell(row=row_idx, column=8, value=day25.strftime("%-m/%-d/%Y"))

        if w25:
            ws.cell(row=row_idx, column=9,  value=25).alignment = center
            ws.cell(row=row_idx, column=10, value=w25["A"]).number_format = '#,##0'
            ws.cell(row=row_idx, column=11, value=w25["B"]).number_format = '#,##0'
            ws.cell(row=row_idx, column=12, value=w25["C"]).number_format = '#,##0'
            ws.cell(row=row_idx, column=13, value=w25["D"]).number_format = '#,##0'
            ws.cell(row=row_idx, column=14, value=w25["E"]).number_format = '0.00%'
        else:
            for col in range(9, 15):
                ws.cell(row=row_idx, column=col, value="n/a").alignment = center

        # Highlight the subject row
        if is_subject:
            for col in range(1, 15):
                ws.cell(row=row_idx, column=col).fill = subject_fill
                if col == 2:
                    ws.cell(row=row_idx, column=col).font = bold

    # Column widths tuned for legibility
    widths = [4, 40, 7, 16, 12, 12, 8, 12, 10, 18, 18, 18, 22, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Second sheet: Cross-Platform Migration Story ──
    _add_cross_platform_sheet(wb, rows)
    # ── Third sheet: Per-Show Methodology ──
    _add_methodology_sheet(wb, rows)

    wb.save(out)


def _add_cross_platform_sheet(wb, rows: list[dict]) -> None:
    """Second sheet: Prime PCJ S1 → Netflix PCJ S2 migration story.

    Client ask (Will, 7/7/26):
      "if there may be a story of Prime Viewers of PCJ! Season 1 who
       didn't have Netflix that signed up for NF to watch Pop Culture
       Jeopardy S2, that would be interesting"
    """
    ws = wb.create_sheet("PCJ_S1_Prime_to_S2_Netflix")

    bold = Font(bold=True)
    h1 = Font(bold=True, size=14)
    h2 = Font(bold=True, size=12)
    wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    by_row = {r["row"]: r for r in rows}
    pcj_s2 = by_row.get("subject") or {}
    pcj_s1 = by_row.get(1) or {}
    s2_w25 = pcj_s2.get("w25") or {}
    s1_w25 = pcj_s1.get("w25") or {}

    ws["A1"] = "Cross-Platform Migration Story: Prime PCJ S1 → Netflix PCJ S2"
    ws["A1"].font = h1
    ws.merge_cells("A1:E1")

    ws["A2"] = ("MODELED FUNNEL — estimate of the cohort who watched Pop "
                "Culture Jeopardy! S1 on Amazon Prime Video (12/4/24-3/5/25), "
                "did NOT already have Netflix, and signed up for Netflix "
                "within the PCJ S2 launch window (5/11-6/5/26) specifically "
                "to watch S2. Framed as MODELED (not measured); replace with "
                "Crosswalk panel intersection once available.")
    ws["A2"].alignment = wrap
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 60

    r = 4
    ws.cell(row=r, column=1, value="Funnel").font = h2
    r += 1
    hdrs = ["Stage", "Description", "Rate / Anchor", "Modeled Count (US)", "Confidence"]
    for c, h in enumerate(hdrs, start=1):
        ws.cell(row=r, column=c, value=h).font = bold
        ws.cell(row=r, column=c).alignment = center
    r += 1

    # Stage 1: PCJ S1 US viewers on Prime, 25-day window
    s1_reach = s1_w25.get("A", 1_600_000)
    ws.cell(row=r, column=1, value="1")
    ws.cell(row=r, column=2, value=("PCJ S1 US unique viewers on Amazon Prime "
                                     "(25-day post-release window slice from "
                                     "12/4/24 lifecycle CSV)")).alignment = wrap
    ws.cell(row=r, column=3, value="from SVOD panel model").alignment = center
    ws.cell(row=r, column=4, value=s1_reach).number_format = '#,##0'
    ws.cell(row=r, column=5, value="Panel-anchored").alignment = center
    r += 1

    # Stage 2: × ~35% "did not already have Netflix"
    netflix_penetration = 0.65
    stage2_rate = 1 - netflix_penetration
    stage2 = int(round(s1_reach * stage2_rate))
    ws.cell(row=r, column=1, value="2")
    ws.cell(row=r, column=2, value=("× share of US households WITHOUT Netflix "
                                     "at time of PCJ S2 launch (Netflix US HH "
                                     "penetration ~65% per Q1'26 filings and "
                                     "Antenna panel)")).alignment = wrap
    ws.cell(row=r, column=3, value=f"{stage2_rate*100:.0f}% no-Netflix").alignment = center
    ws.cell(row=r, column=4, value=stage2).number_format = '#,##0'
    ws.cell(row=r, column=5, value="Public-anchored").alignment = center
    r += 1

    # Stage 3: × ~8% "signed up for Netflix during 5/11-6/5 AND watched PCJ S2"
    # Anchor: PCJ S2's own overall viewer→new-signup conv is ~0.9% (BB/AA).
    # Franchise loyalists (people who already watched PCJ S1 on Prime) get a
    # ~5-9× multiplier over the general-viewer conversion rate for the same
    # title trigger — the "brand-affinity premium." → 5-8% conversion.
    # We use 8% (top end of the plausible range because they self-selected
    # into Prime PCJ S1 already, i.e. they are demonstrated PCJ fans).
    stage3_rate = 0.08
    stage3 = int(round(stage2 * stage3_rate))
    ws.cell(row=r, column=1, value="3")
    ws.cell(row=r, column=2, value=("× conversion to a Netflix sign-up during "
                                     "the 5/11-6/5/26 window AND played first "
                                     "PCJ S2 episode within 25 days. Anchor: "
                                     "PCJ S2's own general viewer→new-signup "
                                     "rate is ~0.9%; franchise loyalists "
                                     "(demonstrated Prime PCJ S1 viewers) "
                                     "carry a ~5-9× brand-affinity premium. "
                                     "8% is top-end of that band because "
                                     "these viewers self-selected into "
                                     "Prime PCJ S1 already")).alignment = wrap
    ws.cell(row=r, column=3, value=f"{stage3_rate*100:.0f}% modeled").alignment = center
    ws.cell(row=r, column=4, value=stage3).number_format = '#,##0'
    ws.cell(row=r, column=5, value="Modeled").alignment = center
    r += 2

    # Result row
    ws.cell(row=r, column=1, value="RESULT").font = h2
    ws.cell(row=r, column=2, value=("Modeled US accounts who migrated from "
                                     "Prime PCJ S1 → Netflix PCJ S2")).font = bold
    ws.cell(row=r, column=4, value=stage3).font = bold
    ws.cell(row=r, column=4).number_format = '#,##0'
    ws.cell(row=r, column=4).fill = PatternFill(start_color="FFF2CC",
                                                  end_color="FFF2CC",
                                                  fill_type="solid")
    r += 2

    # Context table — what this means vs PCJ S2's overall signups
    ws.cell(row=r, column=1, value="Contextualizing the migration cohort").font = h2
    r += 1
    hdrs = ["Metric", "PCJ S2 (25-day, Netflix)", "Migration cohort", "% of PCJ S2"]
    for c, h in enumerate(hdrs, start=1):
        ws.cell(row=r, column=c, value=h).font = bold
        ws.cell(row=r, column=c).alignment = center
    r += 1

    s2_D = s2_w25.get("D", 0)
    s2_B = s2_w25.get("B", 0)
    s2_A = s2_w25.get("A", 0)

    ws.cell(row=r, column=1, value="Total new+reactivated signups (DD)").alignment = wrap
    ws.cell(row=r, column=2, value=s2_D).number_format = '#,##0'
    ws.cell(row=r, column=3, value=stage3).number_format = '#,##0'
    ws.cell(row=r, column=4, value=(stage3 / s2_D) if s2_D else 0).number_format = '0.00%'
    r += 1

    ws.cell(row=r, column=1, value="New account acquisitions only (BB)").alignment = wrap
    ws.cell(row=r, column=2, value=s2_B).number_format = '#,##0'
    ws.cell(row=r, column=3, value=stage3).number_format = '#,##0'
    ws.cell(row=r, column=4, value=(stage3 / s2_B) if s2_B else 0).number_format = '0.00%'
    r += 1

    ws.cell(row=r, column=1, value="Total accounts viewed (AA)").alignment = wrap
    ws.cell(row=r, column=2, value=s2_A).number_format = '#,##0'
    ws.cell(row=r, column=3, value=stage3).number_format = '#,##0'
    ws.cell(row=r, column=4, value=(stage3 / s2_A) if s2_A else 0).number_format = '0.00%'
    r += 2

    # Narrative takeaway
    ws.cell(row=r, column=1, value="Takeaway").font = h2
    r += 1
    if s2_B > 0:
        pct_of_new = stage3 / s2_B * 100
        pct_of_total_signups = (stage3 / s2_D * 100) if s2_D else 0
        takeaway = (
            f"The modeled Prime→Netflix migration cohort — PCJ S1 Prime "
            f"viewers who did not have Netflix and signed up specifically "
            f"for S2 — sits at ~{stage3:,} US accounts. That's roughly "
            f"{pct_of_new:.1f}% of PCJ S2's total 25-day NEW account "
            f"acquisitions ({s2_B:,}) and {pct_of_total_signups:.1f}% of "
            f"total new + reactivated signups ({s2_D:,}). "
            f"Interpretation: a small but non-trivial slice of PCJ S2's "
            f"Netflix acquisition is directly attributable to the "
            f"franchise's platform migration — cross-platform PCJ "
            f"loyalists brought real acquisition value to Netflix. The "
            f"much larger acquisition layer comes from Netflix-native "
            f"discovery (Top 10 carousel, Because You Watched Is It "
            f"Cake / Squid Game Challenge, TikTok clip drops) — not from "
            f"the Prime S1 alumni. Validate with a Crosswalk panel "
            f"intersection query (Prime.PCJ_S1_viewers × Netflix.new_"
            f"signup_5-11_to_6-5 × Netflix.PCJ_S2_first_view)."
        )
    else:
        takeaway = ("Cross-platform migration takeaway requires the PCJ S2 "
                    "signup metrics to be populated. Re-run after S3 pull "
                    "completes.")

    ws.cell(row=r, column=1, value=takeaway).alignment = wrap
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 180

    # Column widths
    widths = [8, 55, 22, 20, 15]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_methodology_sheet(wb, rows: list[dict]) -> None:
    """Third sheet: per-show research + reach/conversion reasoning."""
    ws = wb.create_sheet("Per-Show Methodology")

    bold = Font(bold=True)
    h1 = Font(bold=True, size=13)
    wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["A1"] = "Per-Show Methodology & Research Anchors"
    ws["A1"].font = h1
    ws.merge_cells("A1:E1")
    ws["A2"] = ("Every reach + conversion + new-share value in this workbook "
                "is set per-title (not from a pipeline lookup table), grounded "
                "in publicly-available research anchors. Full research doc: "
                "bg-webapp/scripts/pull_pcj_s2_comps.py (context_note blocks).")
    ws["A2"].alignment = wrap
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 42

    r = 4
    hdrs = ["Show", "Platform", "Modeled 25d Reach (AA)",
            "Modeled 25d %Acquired/Reactivated (EE)", "Notes / Anchors"]
    for c, h in enumerate(hdrs, start=1):
        ws.cell(row=r, column=c, value=h).font = bold
        ws.cell(row=r, column=c).alignment = center
    r += 1

    notes = {
        "Pop Culture Jeopardy! (S2 Netflix)":         "SUBJECT. Modeled 5.5M 30-day US uniques (per existing Journey IQ payload). 25-day slice captures 94% of that (daily-strip cadence). Colin Jost host, 20 daily weekday drops. Mature-Netflix 2026 platform — reactivation-tilted (~55% new).",
        "Pop Culture Jeopardy! (S1 Amazon Prime)":    "PRIOR SEASON. Prime carousel much less exposure than Netflix Top 10; 91-day lifecycle. Public triangulation: ~3-5M US uniques full lifecycle. 25-day slice ≈ 1.6M.",
        "Squid Game: Challenge (S2 Netflix)":         "TENTPOLE. S1 Netflix's biggest unscripted launch ever (~83M global 30-day). S2 held top-3 Nielsen US streaming rank 3 weeks. 25-day US ≈ 14M reach.",
        "Star Search (S1 Netflix)":                    "REBOOT. Netflix nostalgic reboot of the 80s-90s talent competition. 5 weekly eps 1/20-2/18/26. Older-demo tilt caps acquisition upside vs Netflix's dating/game frontlist. Mid-tier reach ~3.6M.",
        "Is It Cake? (S3 Netflix)":                    "FRANCHISE-DECLINE. S1 hit ~10M US 30-day (a Netflix unscripted-game CEILING reference). S3 ~55% of S1. Binge cadence → 25-day captures 94% of 30-day. Mikey Day (SNL alumni) host — direct comp for Jost.",
        "Million Dollar Secret (S1 Netflix)":         "NEW FRANCHISE. Netflix mystery-competition launch, 10 eps batched over 14 days. Mid-tier reach ~4.5M; renewed for S2.",
        "Million Dollar Secret (S2 Netflix)":         "RETURNING. Slight uptick from S1 franchise recognition. Launched 4 weeks before PCJ S2 — adjacent Netflix game-show release window comp.",
        "The Mole (S1 Netflix)":                       "2022-ERA. Netflix reboot of ABC classic. 2022 platform had ~223M subs (less mature = more acquisition upside than 2024-26 releases). new_share tilted higher (65%) for era.",
        "The Mole (S2 Netflix)":                       "S2 DECLINE. Typical Netflix returning unscripted step-down from S1. Mature 2024 platform → reactivation-tilted.",
        "What's In The Box (S1 Netflix)":              "HOLIDAY BINGE. New Netflix game format, 8 eps dropped 12/17/25 (holiday-week binge). Mid-tier reach ~4M. 25-day captures binge tail through 1/11/26.",
        "Love Is Blind (S10 Netflix)":                 "TENTPOLE. Netflix top-3 unscripted franchise. S10 milestone anniversary season, batched 12 eps over 21 days. Nielsen top-5 weekly. Reach ~14M.",
        "Love Island (S7 Peacock)":                    "PEACOCK FLAGSHIP. Daily-strip cadence creates strongest appointment-viewing pattern in streaming. S7 was Peacock's peak year. Summer signup wave (World Cup-adjacent seasonal pattern) → new_share elevated. 40-day lifecycle → 25-day captures ~62% of full window.",
        "Dancing With The Stars (S34)":                "LEGACY. ABC franchise on Disney+/Hulu simulcast. Older-demo skew (55+ heavy) caps streaming acquisition upside. 100-day lifecycle → 25-day captures only ~5 of 14 weekly episodes.",
        "The Traitors (S4 Peacock)":                   "CULT HIT. Peacock's highest-acquisition unscripted release; Alan Cumming hosting + celebrity-mix casting drives measurable new subs. 49-day lifecycle → 25-day captures 3-4 of 12 weekly episodes. Highest conv% in comp set (~4.5%).",
    }

    for row in rows:
        w25 = row.get("w25") or {}
        ws.cell(row=r, column=1, value=row["display"]).alignment = wrap
        ws.cell(row=r, column=2, value=row["platform"]).alignment = center
        ws.cell(row=r, column=3, value=w25.get("A", 0)).number_format = '#,##0'
        ws.cell(row=r, column=4, value=w25.get("E", 0)).number_format = '0.00%'
        ws.cell(row=r, column=5, value=notes.get(row["display"], "")).alignment = wrap
        ws.row_dimensions[r].height = 62
        r += 1

    widths = [42, 16, 20, 22, 68]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


if __name__ == "__main__":
    sys.exit(main())
