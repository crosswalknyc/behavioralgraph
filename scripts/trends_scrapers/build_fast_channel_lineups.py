"""
Build the FAST channel-lineup snapshot from MediaBiz "Stream Metric
Schedules" xlsx files (one per platform).

MediaBiz ships four workbooks, one per FAST platform (Roku Live TV,
Tubi Live TV, Pluto TV Live TV, Amazon Live TV). Each row is a single
airing:

    Country | Service Name | Channel Name | Provider Channel Name
    | Channel Content Type | Current to | Air Date | Start Time
    | End Time | Is Matched | Content Type

The channels within each service are the FAST micro-channels users
actually see on the platform's live-TV grid (e.g. Waypoint TV,
Mr. Bean: Animated, Nick Jr. Pluto TV, Forensic Files 24/7). We
aggregate airings per channel to get a programming-activity signal:
channels with 400+ airings/week are the platform's flagship rails,
channels with a handful of airings are placeholders / launch-week
seeds.

The output snapshot lives at
    s3://dashboard-inputs/trends_iq_snapshots/latest/fast_channel_lineups.json
and is consumed by `trends_iq._fetch_fast_trending`, which attaches a
`channels` list to each platform in the payload.

Usage:
    # default: reads xlsx files from Jenna's Mail Downloads directory
    python3 -m scripts.trends_scrapers.build_fast_channel_lineups

    # explicit paths:
    python3 -m scripts.trends_scrapers.build_fast_channel_lineups \\
        --amazon /path/to/Amazon.xlsx \\
        --roku   /path/to/Roku.xlsx   \\
        --pluto  /path/to/Pluto.xlsx  \\
        --tubi   /path/to/Tubi.xlsx

Refresh cadence: MediaBiz ships these on an ad-hoc basis (currently a
one-time snapshot Jul-Aug 2026). When Jenna receives a new dump,
re-run this script with the new xlsx paths and the snapshot updates
in place - the FAST panel picks up the new lineup on the next
compute_view cache miss.
"""

from __future__ import annotations

import argparse
import glob
import json
import logging
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from typing import Any

logger = logging.getLogger(__name__)


# Slug -> (display label, canonical Service Name value in the xlsx,
#          default Mail-Downloads glob).
# The service name filter matters because a single MediaBiz workbook
# occasionally contains rows tagged for the wrong platform; gate on the
# expected Service Name to keep the snapshot clean.
DEFAULT_MAIL_DIR = os.path.expanduser(
    '~/Library/Containers/com.apple.mail/Data/Library/Mail Downloads')

PLATFORMS = [
    ('amazon', 'Amazon Live TV', 'Stream Metric Schedules Amazon.xlsx'),
    ('roku',   'Roku Live TV',   'Stream Metric Schedules Roku.xlsx'),
    ('pluto',  'Pluto TV Live TV', 'Stream Metric Schedules Pluto.xlsx'),
    ('tubi',   'Tubi Live TV',   'Stream Metric Schedules Tubi.xlsx'),
]

# S3 destination. Mirrors the trends_iq snapshot layout so
# `_read_snapshot('fast_channel_lineups')` picks it up transparently.
S3_BUCKET = 'dashboard-inputs'
S3_KEY_LATEST = 'trends_iq_snapshots/latest/fast_channel_lineups.json'
S3_KEY_DATED_FMT = 'trends_iq_snapshots/{date}/fast_channel_lineups.json'


def _find_default_xlsx(basename: str) -> str | None:
    """Mail.app stores each attachment in its own random-UUID subfolder
    under `~/Library/.../Mail Downloads/`. Glob for the basename across
    all subfolders and return the newest match."""
    if not os.path.isdir(DEFAULT_MAIL_DIR):
        return None
    matches = glob.glob(os.path.join(DEFAULT_MAIL_DIR, '*', basename))
    if not matches:
        return None
    matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return matches[0]


def _process_xlsx(path: str, expected_service: str) -> dict[str, Any]:
    """Parse a MediaBiz Stream Metric Schedules xlsx and aggregate per
    channel. Returns:
        {
            'service':       str,      # canonical service label
            'total_airings': int,      # summed across channels
            'channels':      [ {name, airings, content_type,
                                shows: [], ...}, ... ]  # sorted desc
        }
    """
    import openpyxl

    logger.info("processing %s (%.1f MB)", path,
                 os.path.getsize(path) / 1024 / 1024)

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    airings           = Counter()
    channel_service   = {}
    content_by_ch     = defaultdict(Counter)   # channel -> content_type -> count

    # Header sits at row 9 (0-indexed 8). Data starts at row 10.
    # Column indexes (0-based):
    #   [1]=Country [2]=Service [3]=Channel [5]=Channel Content Type
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 9:
            continue
        try:
            country = row[1]
            service = row[2]
            channel = row[3]
            ch_type = row[5]
        except IndexError:
            continue
        if country and str(country).strip().upper() != 'US':
            continue
        if not channel or not isinstance(channel, str):
            continue
        channel = channel.strip()
        if not channel:
            continue
        # Reject rows whose Service doesn't match this workbook. Some
        # MediaBiz dumps mix rows across services; we want per-platform
        # cleanliness.
        if service and expected_service and \
                str(service).strip().lower() != expected_service.lower():
            continue
        airings[channel] += 1
        if channel not in channel_service and service:
            channel_service[channel] = str(service).strip()
        if ch_type and isinstance(ch_type, str):
            content_by_ch[channel][ch_type.strip()] += 1

    total_airings = sum(airings.values())

    channels = []
    for name, n in airings.most_common():
        # Pick the most-frequent content_type label for the channel
        ct_counter = content_by_ch.get(name) or Counter()
        content_type = ct_counter.most_common(1)[0][0] if ct_counter else ''
        channels.append({
            'name':         name,
            'airings':      n,
            'content_type': content_type,
        })

    logger.info("  %s: %d channels, %s airings",
                 expected_service, len(channels), f'{total_airings:,}')
    return {
        'service':       expected_service,
        'total_airings': total_airings,
        'channels':      channels,
    }


def _upload_to_s3(payload: dict) -> None:
    import boto3
    s3 = boto3.client('s3', region_name='us-east-2')
    body = json.dumps(payload).encode('utf-8')
    today = datetime.now(timezone.utc).date().isoformat()

    s3.put_object(Bucket=S3_BUCKET, Key=S3_KEY_LATEST,
                   Body=body, ContentType='application/json')
    logger.info("wrote s3://%s/%s (%d bytes)",
                 S3_BUCKET, S3_KEY_LATEST, len(body))

    dated = S3_KEY_DATED_FMT.format(date=today)
    s3.put_object(Bucket=S3_BUCKET, Key=dated,
                   Body=body, ContentType='application/json')
    logger.info("wrote s3://%s/%s (%d bytes)", S3_BUCKET, dated, len(body))


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser()
    p.add_argument('--amazon')
    p.add_argument('--roku')
    p.add_argument('--pluto')
    p.add_argument('--tubi')
    p.add_argument('--no-upload', action='store_true',
                    help="skip S3 upload (write to /tmp/ only)")
    p.add_argument('--out',
                    default='/tmp/fast_channel_lineups.json',
                    help="local mirror path (for inspection)")
    args = p.parse_args(argv)

    logging.basicConfig(level=logging.INFO,
                         format='%(asctime)s %(levelname)s %(message)s')

    per_platform: dict[str, dict] = {}
    for slug, service, default_basename in PLATFORMS:
        override = getattr(args, slug, None)
        path = override or _find_default_xlsx(default_basename)
        if not path:
            logger.warning("skipping %s: no xlsx found (looked for %s in %s)",
                            slug, default_basename, DEFAULT_MAIL_DIR)
            continue
        if not os.path.exists(path):
            logger.warning("skipping %s: file does not exist: %s", slug, path)
            continue
        try:
            per_platform[slug] = _process_xlsx(path, service)
        except Exception as e:
            logger.exception("failed to process %s (%s): %s", slug, path, e)

    if not per_platform:
        logger.error("no platforms processed; nothing to write")
        return 2

    payload = {
        'source':      'fast_channel_lineups',
        'label':       'FAST channel lineups',
        'kind':        'fast_channels',
        'fetched_at':  datetime.now(timezone.utc).isoformat(),
        'sources':     per_platform,
    }

    with open(args.out, 'w') as f:
        json.dump(payload, f)
    logger.info("wrote local mirror: %s (%d bytes)",
                 args.out, os.path.getsize(args.out))

    if not args.no_upload:
        try:
            _upload_to_s3(payload)
        except Exception as e:
            logger.exception("S3 upload failed: %s", e)
            return 3
    else:
        logger.info("--no-upload set; skipping S3")

    # Summary line
    for slug, data in per_platform.items():
        n = len(data.get('channels') or [])
        top5 = ', '.join(c['name'] for c in (data.get('channels') or [])[:5])
        print(f"{slug:8s}  {data.get('service','?'):20s}  "
               f"{n:4d} channels  top5=[{top5}]")

    return 0


if __name__ == '__main__':
    sys.exit(main())
